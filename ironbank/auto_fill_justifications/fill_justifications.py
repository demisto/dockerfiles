import argparse
from pathlib import Path
from openpyxl import load_workbook
from justifications_mapping import SHEET_JUSTIFICATIONS_MAPPERS

MAX_OFFSET_RETRIES = 3
WARNING_STYLE = '\033[93m'
OK_GREEN_STYLE ='\033[92m'
END_COLOR = '\033[0m'
SHEET_YELLOW_COLORS = ['00ffff00', 'FFFFFF00']


def args_handler():
    parser = argparse.ArgumentParser(description='Fill justification file automatically.')
    parser.add_argument('--justification-path', help='The justification file path.', required=True)
    return parser.parse_args()


def find_indexes(sheet, justification_col_name):
    offset = 1
    justification_index = 0
    justification_id_index = 0

    # some sheet headers didn't starts from the first row(index 1), if we couldn't gets the indexes in the first row,
    # the while statement will look for it in the next row.
    while offset < MAX_OFFSET_RETRIES and not justification_id_index or not justification_index:
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=offset, column=column).value

            # try to get the justification and justification id indexes.
            if cell_value == 'Justification':
                justification_index = column
            if cell_value == justification_col_name:
                justification_id_index = column

            if justification_id_index and justification_index:
                return justification_id_index, justification_index

        offset = offset + 1

    return justification_id_index, justification_index


def fill_justifications(sheet, justifications, trigger_id_index, justification_index):
    warnings = set()

    for row in range(1, sheet.max_row + 1):
        justification_cell = sheet.cell(row=row, column=justification_index)

        # if the justification column background is yellow, fill it with the value from the mapper.
        if justification_cell.fill.bgColor.index in SHEET_YELLOW_COLORS:
            trigger_id_value = sheet.cell(row=row, column=trigger_id_index).value
            if justifications.get(trigger_id_value):
                justification_cell.value = justifications[trigger_id_value]
            else:
                warnings.add(f'No justification found for {trigger_id_value} in the sheet: {sheet.title}')

    # print all the warnings
    for warn in warnings:
        print(f'{WARNING_STYLE}Warning: {warn}{END_COLOR}')


def main():
    args = args_handler()
    input_file = args.justification_path

    p = Path(input_file)
    # check if the justification file exists
    if not p.is_file():
        raise Exception(f'Could not find the justification file in {input_file}')

    # set the result file name
    res_file = str(Path(p.parent, f"{p.stem}_FILLED{p.suffix}"))

    print(f'Processing the file {input_file}...\n')

    workbook = load_workbook(filename=input_file)

    for sheet in workbook.worksheets:
        # check if we have mapper for current sheet
        if sheet.title in SHEET_JUSTIFICATIONS_MAPPERS.keys():
            mapper = SHEET_JUSTIFICATIONS_MAPPERS[sheet.title]
            justification_col_name = mapper['column_name']

            # get the columns index for justification and justification id
            justification_id_index, justification_index = find_indexes(sheet, justification_col_name)

            if not justification_id_index or not justification_index:
                raise Exception(f'Failed to find the column "{justification_col_name}" in the sheet: {sheet.title}')

            justifications = mapper['justifications']
            # fill the justifications in current sheet
            fill_justifications(sheet, justifications, justification_id_index, justification_index)

    # save the file
    workbook.save(res_file)
    print(f'{OK_GREEN_STYLE}{res_file} created successfully!{END_COLOR}')


if __name__ == '__main__':
    main()

